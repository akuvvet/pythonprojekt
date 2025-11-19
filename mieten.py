import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.cell.cell import MergedCell
import os
import re
from pandas.api.types import is_datetime64_any_dtype, is_numeric_dtype
from datetime import datetime

MONATS_ZUORDNUNG = {
    # Hinweis: Wegen neuer Spalte "Objekt" zwischen B und C sind alle Zielspalten +1 verschoben
    "Jan": ["E", "F"],
    "Feb": ["G", "H"],
    "Mrz": ["I", "J"],
    "Apr": ["K", "L"],
    "Mai": ["M", "N"],
    "Jun": ["O", "P"],
    "Jul": ["Q", "R"],
    "Aug": ["S", "T"],
    "Sep": ["U", "V"],
    "Okt": ["W", "X"],
    "Nov": ["Y", "Z"],
    "Dez": ["AA", "AB"],
}

MONATS_NAMENS_MAPPING = {
    "Januar": "Jan", "Jan": "Jan",
    "Februar": "Feb", "Feb": "Feb",
    "März": "Mrz", "Marz": "Mrz", "Mrz": "Mrz",
    "April": "Apr", "Apr": "Apr",
    "Mai": "Mai",
    "Juni": "Jun", "Jun": "Jun",
    "Juli": "Jul", "Jul": "Jul",
    "August": "Aug", "Aug": "Aug",
    "September": "Sep", "Sep": "Sep", "Sept": "Sep",
    "Oktober": "Okt", "Okt": "Okt",
    "November": "Nov", "Nov": "Nov",
    "Dezember": "Dez", "Dez": "Dez",
}

BLATTNAME = "mieter"
MIETER_SPALTE = "A"

# Kontoauszug (XLSX) erwartete Spalten
KONTO_DATUM = "Wertstellung"
KONTO_PAYEE = "Empfänger/Auftraggeber"
KONTO_VWZ = "Verwendungszweck"
KONTO_KATEGORIE = "Kategorie"
KONTO_OBJEKT = "Kontoname (Objekt)"
KONTO_BETRAG = "Betrag"


def fuehre_mietabgleich_durch(excel_pfad, konto_xlsx_pfad):

    # Excel (Mieterliste) einlesen
    workbook = load_workbook(excel_pfad)
    if BLATTNAME in workbook.sheetnames:
        try:
            worksheet = workbook[BLATTNAME]
        except Exception:
            return None
        sheet_arg = BLATTNAME
    else:
        # Fallback: erstes Blatt
        first_sheet = workbook.sheetnames[0]
        worksheet = workbook[first_sheet]
        sheet_arg = first_sheet

    # Mieterliste laden
    df_mieter = pd.read_excel(excel_pfad, sheet_name=sheet_arg, dtype=str)
    df_mieter = df_mieter.fillna("")
    mieter_col_name = df_mieter.columns[0]  # Spalte A: Eigentümer/Mieter
    mieter_b_col_name = df_mieter.columns[1] if len(df_mieter.columns) > 1 else None  # Spalte B: Mieter
    objekt_col_name = df_mieter.columns[2] if len(df_mieter.columns) > 2 else None     # Spalte C: Objekt

    # Erzeuge ein Index-Mapping von Namen (Spalte A) -> Zeilennummer im Zielblatt,
    # damit Einträge immer in die korrekte Zeile geschrieben werden – unabhängig von Einfügungen.
    mieter_row_map = {}
    try:
        for r in range(1, worksheet.max_row + 1):
            cell_val = worksheet[f"{MIETER_SPALTE}{r}"].value
            if cell_val is None:
                continue
            key = str(cell_val)
            # gleiche Normalisierung wie beim Matching
            key = re.sub(r"[^a-z0-9\s]", " ", key.lower())
            key = key.replace("ä","ae").replace("ö","oe").replace("ü","ue").replace("ß","ss")
            key = re.sub(r"\s+", " ", key).strip()
            if key and key not in mieter_row_map:
                mieter_row_map[key] = r
    except Exception:
        pass

    # Kontoauszug (XLSX) einlesen
    try:
        df_konto = pd.read_excel(konto_xlsx_pfad, dtype=str)
    except Exception:
        return None

    df_konto = df_konto.fillna("")

    # Fallback für den Fall, dass keine Header vorhanden sind (A-F)
    if not all(col in df_konto.columns for col in [KONTO_DATUM, KONTO_PAYEE, KONTO_VWZ, KONTO_KATEGORIE, KONTO_OBJEKT, KONTO_BETRAG]):
        if len(df_konto.columns) >= 6:
            cols = list(df_konto.columns[:6])
            mapping = {
                cols[0]: KONTO_DATUM,
                cols[1]: KONTO_PAYEE,
                cols[2]: KONTO_VWZ,
                cols[3]: KONTO_KATEGORIE,
                cols[4]: KONTO_OBJEKT,
                cols[5]: KONTO_BETRAG,
            }
            df_konto = df_konto.rename(columns=mapping)

    # Typ-Konvertierungen (robust für EU-Formate wie 640,80 und Tausenderpunkte)
    df_konto["__betrag_raw"] = df_konto[KONTO_BETRAG].astype(str)

    def _parse_amount_raw(raw: str):
        s = (str(raw) or "").strip()
        if not s:
            return pd.NA
        s = s.replace("€", "").replace("\xa0", "").replace(" ", "")
        # Wenn Komma vorhanden → als Dezimaltrenner behandeln, Punkte als Tausender entfernen
        if "," in s:
            s2 = s.replace(".", "").replace(",", ".")
            try:
                return float(s2)
            except Exception:
                pass
        # Sonst direkten Float-Versuch (z. B. 640.80)
        try:
            return float(s)
        except Exception:
            # Letzter Fallback: Muster <ganzzahl><,|.><1-2 Dezimalstellen>
            m = re.search(r"(\d+)[,\.](\d{1,2})$", s)
            if m:
                try:
                    return float(m.group(1) + "." + m.group(2))
                except Exception:
                    pass
        return pd.NA

    df_konto[KONTO_BETRAG] = df_konto["__betrag_raw"].apply(_parse_amount_raw)

    # Datumsformat sauber parsen und Rohwert behalten
    s = df_konto[KONTO_DATUM]
    if is_datetime64_any_dtype(s):
        df_konto["__raw_date"] = s.dt.strftime("%d.%m.%Y")
    elif is_numeric_dtype(s):
        dt = pd.to_datetime(s, unit="d", origin="1899-12-30", errors="coerce")
        df_konto["__raw_date"] = dt.dt.strftime("%d.%m.%Y")
        df_konto[KONTO_DATUM] = dt
    else:
        s2 = (
            s.astype(str)
             .str.strip()
             .str.replace(r"\s+", "", regex=True)
             .str.replace("/", ".", regex=False)
        )
        df_konto["__raw_date"] = s2
        df_konto[KONTO_DATUM] = pd.to_datetime(s2, format="%d.%m.%Y", errors="coerce")

    # Zahlungsgrund klassifizieren (Miete > Nebenkosten > Nachzahlung > Rate > Honorar)
    def klassifiziere(text: str) -> str:
        t = (text or "").lower()
        # Robustere Erkennung inkl. Synonyme/Abkürzungen; Reihenfolge = Priorität
        patterns = [
            ("Miete", [
                r"\bmiet\w*\b",  # deckt miete, mieten, mietzahlung, mietzins, mieter, etc. ab
                r"\bkm\b", r"\bkaltmiete\b",
                r"\bstellplatz\b", r"\bgarage\b"
            ]),
            ("Nebenkosten", [
                r"\bnebenkosten\b", r"\bnk\b", r"\bbetriebskosten\b", r"\bbk\b"
            ]),
            ("Nachzahlung", [
                r"\bnach\-?zahlung\b", r"\bnachz\b"
            ]),
            ("Rate", [
                r"\brate(nzahlung)?\b"
            ]),
            ("Honorar", [
                r"\bhonorar\b"
            ]),
        ]
        for label, pats in patterns:
            for p in pats:
                if re.search(p, t, re.IGNORECASE):
                    return label
        return "Sonstiges"

    df_konto["__text_summe"] = (
        df_konto[KONTO_VWZ].astype(str) + " " +
        df_konto[KONTO_KATEGORIE].astype(str) + " " +
        df_konto[KONTO_OBJEKT].astype(str)
    )
    df_konto["__klass"] = df_konto["__text_summe"].apply(klassifiziere)

    # Trefferwort (erstes passendes Suchwort) für spätere Auswertung
    def finde_suchwort(text: str) -> str:
        t = (text or "").lower()
        patterns = [
            ("Miete", [
                r"\bmiet\w*\b", r"\bkm\b", r"\bkaltmiete\b",
                r"\bstellplatz\b", r"\bgarage\b"
            ]),
            ("Nebenkosten", [
                r"\bnebenkosten\b", r"\bnk\b", r"\bbetriebskosten\b", r"\bbk\b"
            ]),
            ("Nachzahlung", [
                r"\bnach\-?zahlung\b", r"\bnachz\b"
            ]),
            ("Rate", [
                r"\brate(nzahlung)?\b"
            ]),
            ("Honorar", [
                r"\bhonorar\b"
            ]),
        ]
        for _, pats in patterns:
            for p in pats:
                m = re.search(p, t, re.IGNORECASE)
                if m:
                    return m.group(0)
        return ""

    df_konto["__hit"] = df_konto["__text_summe"].apply(finde_suchwort)

    # Monatsangabe im Verwendungszweck/Kategorie ermitteln (hat Vorrang vor Wertstellung)
    month_key_map = {k.lower(): v for k, v in MONATS_NAMENS_MAPPING.items()}
    month_pattern = re.compile(r"\b(" + "|".join(re.escape(k) for k in MONATS_NAMENS_MAPPING.keys()) + r")\b", flags=re.IGNORECASE)
    def finde_monats_override(vwz_text: str) -> str | None:
        if not vwz_text:
            return None
        m = month_pattern.search(vwz_text)
        if not m:
            return None
        return month_key_map.get(m.group(1).lower())

    df_konto["__month_override"] = (
        (df_konto[KONTO_VWZ].astype(str) + " " + df_konto[KONTO_KATEGORIE].astype(str))
        .apply(finde_monats_override)
    )

    # Sonderfall: Wenn Zahlender eine Behörde ist (Jobcenter/Agentur/Stadt Wuppertal),
    # soll im Blatt "suchtreffer" in der Spalte "Suchwort" der komplette Verwendungszweck stehen.
    def _normalize_simple(val: str) -> str:
        t = (str(val) or "").lower()
        t = t.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
        t = re.sub(r"[^a-z0-9\s]", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t

    payee_norm = df_konto[KONTO_PAYEE].astype(str).apply(_normalize_simple)
    gov_mask = (
        payee_norm.str.contains(r"\bjobcenter\b", regex=True)
        | payee_norm.str.contains(r"\bbundesagentur\b", regex=True)
        | payee_norm.str.contains(r"\bstadt\s*wuppertal\b", regex=True)
    )
    # hit_final: bei Behörden kompletter Verwendungszweck, sonst ermitteltes Suchwort (oder Klasse als Fallback)
    df_konto["__hit_final"] = df_konto[KONTO_VWZ].astype(str)
    df_konto.loc[~gov_mask, "__hit_final"] = df_konto["__hit"]
    df_konto["__hit_final"] = df_konto["__hit_final"].where(df_konto["__hit_final"] != "", df_konto["__klass"])

    # Neues Blatt mit Suchtreffern erstellen: A Datum, B Name, C Suchwort, D Betrag
    sheet_such = "suchtreffer"
    if sheet_such in workbook.sheetnames:
        del workbook[sheet_such]
    ws_such = workbook.create_sheet(sheet_such)
    ws_such.append(["Datum", "Name", "Suchwort", "Betrag"])

    relevante_labels = {"Miete", "Nebenkosten", "Nachzahlung", "Rate", "Honorar"}
    # Alle Einzelbuchungen ohne Aggregation; stabil sortieren für nachvollziehbare Reihenfolge
    df_such = df_konto[df_konto["__klass"].isin(relevante_labels)].copy()
    try:
        df_such = df_such.sort_values([KONTO_PAYEE, KONTO_DATUM, KONTO_BETRAG], kind="mergesort")
    except Exception:
        pass

    for _, r in df_such.iterrows():
        hit = r["__hit_final"]
        ws_such.append(["", r[KONTO_PAYEE], str(hit), r[KONTO_BETRAG]])
        row_idx = ws_such.max_row
        # Datum korrekt schreiben (immer als echtes Datum DD.MM.YYYY; bei Fehler leer lassen)
        date_cell = ws_such.cell(row=row_idx, column=1)
        raw_val = r.get("__raw_date", "")
        dval = r[KONTO_DATUM]
        try:
            if pd.notna(dval):
                py_dt = dval.to_pydatetime() if hasattr(dval, "to_pydatetime") else dval
                date_cell.value = py_dt.date()
                date_cell.number_format = "DD.MM.YYYY"
            else:
                rv = (str(raw_val) or "").strip()
                # ISO-ähnlich: 2025-06-02...
                m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
                if m_iso:
                    parsed = datetime(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
                    date_cell.value = parsed.date()
                    date_cell.number_format = "DD.MM.YYYY"
                else:
                    rv2 = rv.replace("/", ".")
                    parsed = datetime.strptime(rv2, "%d.%m.%Y")
                    date_cell.value = parsed.date()
                    date_cell.number_format = "DD.MM.YYYY"
        except Exception:
            date_cell.value = None
            date_cell.number_format = "DD.MM.YYYY"
        # Betrag formatieren
        try:
            ws_such.cell(row=row_idx, column=4).number_format = "#,##0.00"
        except Exception:
            pass

    def _normalize_text(val: str) -> str:
        t = (str(val) or "").lower()
        t = t.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
        t = re.sub(r"[^a-z0-9\s]", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t

    df_konto["__norm_payee"] = df_konto[KONTO_PAYEE].astype(str).apply(_normalize_text)
    df_konto["__norm_kombi"] = (
        df_konto[KONTO_PAYEE].astype(str) + " " +
        df_konto[KONTO_VWZ].astype(str) + " " +
        df_konto[KONTO_OBJEKT].astype(str)
    ).apply(_normalize_text)
    df_konto["__norm_vwz"] = df_konto[KONTO_VWZ].astype(str).apply(_normalize_text)
    df_konto["__norm_objekt"] = df_konto[KONTO_OBJEKT].astype(str).apply(_normalize_text)
    # Für Behörden-Fall: Suchwort (Verwendungszweck) normalisiert
    try:
        df_such["__norm_hit"] = df_such["__hit_final"].astype(str).apply(_normalize_text)
    except Exception:
        pass

    # Eintragen aus Blatt "suchtreffer" in Monats-Spalten (E–AB) je Mieter (Spalte A)
    months_order = list(MONATS_ZUORDNUNG.keys())

    # Helfer
    def _get_writable_cell(ws, coord):
        c = ws[coord]
        if isinstance(c, MergedCell):
            for r in ws.merged_cells.ranges:
                if coord in r:
                    return ws.cell(row=r.min_row, column=r.min_col)
        return c

    def _write_date(cell, dt_val, raw_val) -> str:
        # Schreibe IMMER als echtes Datum (DD.MM.YYYY); bei Fehler leer
        if pd.notna(dt_val):
            try:
                py_dt = dt_val.to_pydatetime() if hasattr(dt_val, "to_pydatetime") else dt_val
                cell.value = py_dt.date()
                cell.number_format = "DD.MM.YYYY"
                return py_dt.strftime("%d.%m.%Y")
            except Exception:
                pass
        rv = (str(raw_val) or "").strip()
        # ISO-ähnlich: 2025-06-02...
        m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
        try:
            if m_iso:
                parsed = datetime(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
            else:
                parsed = datetime.strptime(rv.replace("/", "."), "%d.%m.%Y")
            cell.value = parsed.date()
            cell.number_format = "DD.MM.YYYY"
            return parsed.strftime("%d.%m.%Y")
        except Exception:
            cell.value = None
            cell.number_format = "DD.MM.YYYY"
            return ""

    def _parse_amount_cell(val) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    def _norm_ddmmyyyy(s: str) -> str:
        s = (s or "").strip()
        m = re.match(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$", s)
        if not m:
            return s
        d = int(m.group(1)); mth = int(m.group(2)); yr = m.group(3)
        if yr is None:
            return f"{d:02d}.{mth:02d}"
        return f"{d:02d}.{mth:02d}.{int(yr):04d}"

    def _parse_pairs(txt: str):
        pairs = []
        seen = set()
        if not txt:
            return pairs
        for line in txt.splitlines():
            m = re.search(r"(\d{1,2}\.\d{1,2}(?:\.\d{2,4})?)\s*(?:\[(.*?)\])?\s*:\s*([+-]?\d+(?:[.,]\d+)?)", line)
            if not m:
                continue
            d = _norm_ddmmyyyy(m.group(1))
            try:
                amt = round(float(m.group(3).replace(".", "").replace(",", ".")), 2)
            except Exception:
                continue
            key = (d, amt)
            if key in seen:
                continue
            seen.add(key)
            pairs.append((d, amt, m.group(1), (m.group(2) or "").strip()))
        return pairs

    # Normname in Trefferliste
    df_such["__norm_payee"] = df_such[KONTO_PAYEE].astype(str).apply(_normalize_text)

    for _, row in df_mieter.iterrows():
        m_name = row[mieter_col_name]
        if not m_name:
            continue
        excel_row = mieter_row_map.get(_normalize_text(m_name))
        if not excel_row:
            continue

        # Behördenfall: Wenn Spalte A "jobcenter"/"agentur"/"stadt wuppertal" enthält,
        # suche in suchtreffer den Namen aus Spalte B (Mieter) als Substring in Name (Spalte B).
        owner_norm = _normalize_text(m_name)
        tenant_norm = _normalize_text(row[mieter_b_col_name]) if mieter_b_col_name else ""
        GOV_KEYS = ("jobcenter", "agentur", "stadt wuppertal")
        if any(k in owner_norm for k in GOV_KEYS) and tenant_norm:
            # Suche den Mieternamen (Spalte B aus "mieter") im Suchwort/Verwendungszweck (Spalte C in "suchtreffer")
            treffer = df_such[df_such.get("__norm_hit", df_such["__norm_payee"]).str.contains(tenant_norm, na=False)]
        else:
            treffer = df_such[df_such["__norm_payee"] == owner_norm]
        if treffer.empty:
            continue

        for _, t in treffer.iterrows():
            dval = t[KONTO_DATUM]
            raw_date = t.get("__raw_date", "")
            kw = t["__hit"] if t["__hit"] else t["__klass"]
            betrag = t[KONTO_BETRAG]
            if pd.isna(betrag):
                continue

            # Monat robust bestimmen (Timestamp | ISO `YYYY-MM-DD...` | `DD.MM.YYYY`)
            month_idx = None
            # 0) Vorrang: Monatsangabe aus Verwendungszweck/Kategorie
            override = t.get("__month_override", None)
            if isinstance(override, str) and override in months_order:
                month_idx = months_order.index(override) + 0  # already 0-based via index
            if pd.notna(dval):
                try:
                    month_idx = int(getattr(dval, "month"))
                except Exception:
                    month_idx = None
            if month_idx is None:
                rv = str(raw_date or "").strip()
                # ISO-ähnlich: 2025-06-02... → 06
                m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
                if m_iso:
                    try:
                        month_idx = int(m_iso.group(2))
                    except Exception:
                        month_idx = None
                if month_idx is None:
                    try:
                        month_idx = datetime.strptime(rv, "%d.%m.%Y").month
                    except Exception:
                        month_idx = None
            if month_idx is None:
                continue
            month_idx = month_idx - 1
            if month_idx < 0 or month_idx > 11:
                continue
            ziel = months_order[month_idx]
            betrag_sp, datum_sp = MONATS_ZUORDNUNG[ziel]
            betrag_cell = f"{betrag_sp}{excel_row}"
            datum_cell = f"{datum_sp}{excel_row}"

            date_cell = _get_writable_cell(worksheet, datum_cell)
            prev_cell_val = date_cell.value
            existing_amount = _parse_amount_cell(worksheet[betrag_cell].value)
            existing_pairs = _parse_pairs(date_cell.comment.text if date_cell.comment else "")
            existing_keys = {(d, a) for (d, a, _, __) in existing_pairs}

            # neuer Schlüssel (Datum+Betrag) – Datum robust formatiert
            if pd.notna(dval):
                try:
                    new_date_str = dval.strftime("%d.%m.%Y")
                except Exception:
                    new_date_str = ""
            else:
                rv = str(raw_date or "").strip()
                m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
                if m_iso:
                    try:
                        new_date_str = datetime(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3))).strftime("%d.%m.%Y")
                    except Exception:
                        new_date_str = rv
                else:
                    try:
                        new_date_str = datetime.strptime(rv, "%d.%m.%Y").strftime("%d.%m.%Y")
                    except Exception:
                        new_date_str = rv
            new_key = (_norm_ddmmyyyy(new_date_str), round(float(betrag), 2))
            # 1) Duplikat prüfen gegen vorhandene Kommentar-Paare
            if new_key in existing_keys:
                continue
            # 2) Duplikat prüfen gegen erste Einzelbuchung ohne Kommentar
            if (existing_amount > 0.0) and not existing_pairs:
                prev_str = prev_cell_val.strftime("%d.%m.%Y") if hasattr(prev_cell_val, "strftime") else (str(prev_cell_val) if prev_cell_val else "")
                prev_key = (_norm_ddmmyyyy(prev_str), round(existing_amount, 2))
                if prev_key == new_key:
                    continue

            # Betrag addieren und Datum setzen
            worksheet[betrag_cell] = existing_amount + float(betrag)
            try:
                worksheet[betrag_cell].number_format = "#,##0.00"
            except Exception:
                pass
            written_date_str = _write_date(date_cell, dval, raw_date)

            # Kommentar nur bei „zweitem+“ Eintrag im selben Monat
            has_previous_in_month = (existing_amount > 0.0) or (len(existing_pairs) > 0)
            if has_previous_in_month:
                # Kommentar pflegen: „DD.MM.YYYY [Suchwort]: 123,45 EUR“
                lines = []
                if existing_pairs:
                    for (_, amt, disp, kword) in existing_pairs:
                        tag = f"{disp}"
                        if kword:
                            tag += f" [{kword}]"
                        lines.append(f"{tag}: {str(f'{amt:.2f}').replace('.', ',')} EUR")
                else:
                    # Es gab bereits einen Betrag, aber noch keinen Kommentar → mit bisherigem Datum/Betrag seeden
                    prev_str = prev_cell_val.strftime("%d.%m.%Y") if hasattr(prev_cell_val, "strftime") else (str(prev_cell_val) if prev_cell_val else "")
                    if prev_str:
                        lines.append(f"{prev_str}: {str(f'{existing_amount:.2f}').replace('.', ',')} EUR")
                tag_new = written_date_str
                if kw:
                    tag_new += f" [{kw}]"
                lines.append(f"{tag_new}: {str(f'{float(betrag):.2f}').replace('.', ',')} EUR")
                date_cell.comment = Comment("\n".join(lines), "System")
            else:
                # Beim ersten Eintrag keinen Kommentar hinterlegen
                date_cell.comment = None

    # Speichern in results/ Ordner
    result_path = os.path.join("results", "mieten_abgleich.xlsx")
    workbook.save(result_path)

    return result_path
